import openpyxl


class HomePageData:
  test_homepagedata = [{"first_name":"Chandani","email":"chandani.jain@gmail.com","password":"12345","gender_index":"1"},{"first_name":"Deepak","email":"deepak@gmail.com","password":"7890","gender_index":"0"}]
  
  @staticmethod
  def get_data(testcasename):
    Dict = {}
    book = openpyxl.load_workbook("C:\\Users\\chand\\OneDrive\\Desktop\\Python Course\\Python Sel\\excelData.xlsx")
    sheet = book["Sheet1"]
    for i in range(1, sheet.max_row+1):
      if sheet.cell(row=i, column=1).value == testcasename:
        for j in range(1, sheet.max_column+1):
          Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
          #print(sheet.cell(row=i, column=j).value)
          
    return[Dict]