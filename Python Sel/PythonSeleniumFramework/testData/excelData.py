from ast import Dict
import openpyxl

Dict = {}
book = openpyxl.load_workbook("C:\\Users\\chand\\OneDrive\\Desktop\\Python Course\\Python Sel\\excelData.xlsx")
sheet = book["Sheet1"]        #can do book.active which gives the active sheet
cell = sheet.cell(row=1, column=2)
print(cell.value)       #reading from a cell

sheet.cell(row=2, column=2).value = "Chandani"      #writing to a cell
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet["B1"].value)      #same as row1 col2

#print all the values in the excel
print("All the values of excel")
for i in range(1, sheet.max_row+1):      #+1 as the last index is exclusive
  for j in range(1, sheet.max_column+1):
    print(sheet.cell(row=i, column=j).value)
    
#get data only for a particular testcase
print("Values from excel for a particular TC: TC2")
for i in range(1, sheet.max_row+1):
  if sheet.cell(row=i, column=1).value == "TC2":
    for j in range(1, sheet.max_column+1):
      Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
      #print(sheet.cell(row=i, column=j).value)
      
print(Dict)